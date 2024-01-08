import requests, json
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from rest_framework import generics, permissions
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework import status
from rest_framework.generics import ListAPIView, RetrieveUpdateAPIView
from .serializers import (
    PhoneNumberSerializer,
    ExcelSerializer,
    ExcelImageSerializer,
    WhatsAppBulkMessageSerializer,
    WhatsAppBulkMessageImageSerializer,
    MessageTemplateSerializer,
    MessageTextTemplateSerializer,
    ImageUploadSerializer,
    CustomUserSerializer,
    CustomUserDetailSerializer,
    CredentialsSerializer,
    UserLoginSerializer,
    ReferalStringSerializer,
    ScheduledAPISerializer,
)
from rest_framework_simplejwt.tokens import RefreshToken
from .models import (
    PhoneNumber,
    CustomUser,
    WhatsappCredential,
    Template,
    ScheduledAPICall,
)
from openpyxl import load_workbook
from decouple import config
from django.contrib.auth import authenticate
import urllib.parse
import base64, os, random, string

my_token = "your_verify_token"
bearer_token = config("TOKEN")
phone_number_id = config("PHONE_NO_ID")
business_id = config("BUSINESS_ID")
domain_url = "http://127.0.0.1:8000/media/"


def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)

    refresh.payload["user_id"] = user.id
    refresh.payload["user_is_staff"] = user.is_staff
    refresh.payload["user_is_distributor"] = user.is_distributor

    return {
        "access": str(refresh.access_token),
    }


def get_credentials(user_id):
    # get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    # business_id = get_credential[0]["whatsapp_business_id"]
    # bearer_token = get_credential[0]["permanent_access_token"]

    try:
        credentials = WhatsappCredential.objects.filter(user_id=user_id).values(
            "phone_number_id", "whatsapp_business_id", "permanent_access_token"
        )
        return credentials
    except WhatsappCredential.DoesNotExist:
        return None


@api_view(["POST"])
# @permission_classes([IsAuthenticated])
def validate_access_token(request):
    user = request.user

    try:
        refresh = RefreshToken.for_user(user)

        refresh.access_token.verify()

        return Response({"valid": True})
    except Exception as e:
        return Response({"valid": False}, status=401)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_credentials(request):
    if request.method == "POST":
        serializer = CredentialsSerializer(data=request.data)
        if serializer.is_valid():
            user_id = serializer.validated_data["user_id"]

            print(phone_number_id, business_id, bearer_token)
            try:
                credentials = WhatsappCredential.objects.get(user_id=user_id)
                print(credentials)
                user = CustomUser.objects.get(id=user_id)

                credentials.whatsapp_business_id = serializer.validated_data[
                    "whatsapp_business_id"
                ]
                credentials.permanent_access_token = serializer.validated_data[
                    "permanent_access_token"
                ]
                credentials.phone_number_id = serializer.validated_data[
                    "phone_number_id"
                ]
                credentials.save()

            except WhatsappCredential.DoesNotExist:
                user = CustomUser.objects.get(id=user_id)

                WhatsappCredential.objects.create(
                    user=user,
                    phone_number_id=serializer.validated_data["phone_number_id"],
                    whatsapp_business_id=serializer.validated_data[
                        "whatsapp_business_id"
                    ],
                    permanent_access_token=serializer.validated_data[
                        "permanent_access_token"
                    ],
                )

            return Response(
                {"message": "Data saved successfully"}, status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserRegistrationView(APIView):
    def post(self, request):
        serializer = CustomUserSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            response_data = {
                "message": "User registered successfully",
                "user": serializer.data,
            }
            return Response(response_data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserLoginView(APIView):
    def post(self, request):
        serializer = UserLoginSerializer(data=request.data)
        if serializer.is_valid():
            username = serializer.validated_data.get("username")
            password = serializer.validated_data.get("password")

            user = serializer.validated_data["user"]
            if user is not None and user.is_active:
                print(user.is_active)
                token = get_tokens_for_user(user)
                return Response(
                    {
                        "token": token,
                        "user_id": user.id,
                        "is_manager": user.is_staff,
                        "is_active": user.is_active,
                        "message": "Login successful",
                    },
                    status=status.HTTP_200_OK,
                )
            else:
                return Response(
                    {"message": "Invalid credentials or inactive account"},
                    status=status.HTTP_401_UNAUTHORIZED,
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserListView2(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request, format=None):
        users = CustomUser.objects.filter(is_staff="False")
        serializer = CustomUserSerializer(users, many=True)
        return Response(serializer.data)


class UserDetailView(RetrieveUpdateAPIView):
    permission_classes = [IsAdminUser]
    queryset = CustomUser.objects.all()
    serializer_class = CustomUserDetailSerializer


class UserChildrenListView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request, *args, **kwargs):
        parent_user_id = self.kwargs.get("pk")

        if parent_user_id:
            children = CustomUser.objects.filter(parent_user_id=parent_user_id)
            data = [
                {"id": child.id, "email": child.email, "is_active": child.is_active}
                for child in children
            ]
            return Response(data)
        else:
            return Response([])


class UserHierarchyView(generics.RetrieveAPIView):
    permission_classes = [permissions.IsAdminUser]

    def retrieve(self, request, *args, **kwargs):
        user_id = self.kwargs.get("pk")
        user = CustomUser.objects.get(pk=user_id)
        hierarchy_data = self.get_user_hierarchy(user)
        return Response(hierarchy_data)

    def get_user_hierarchy(self, user):
        return self.build_hierarchy(user)

    def build_hierarchy(self, user):
        user_data = {
            "id": user.id,
            "email": user.email,
            "is_active": user.is_active,
            "is_distributor": user.is_distributor,
            "children": [],
        }

        children = CustomUser.objects.filter(parent_user=user)
        for child in children:
            user_data["children"].append(self.build_hierarchy(child))

        return user_data

    def get_user_descendants(self, parent_user_id):
        descendants = []
        children = CustomUser.objects.filter(parent_user_id=parent_user_id)

        for child in children:
            descendants.append(child)
            descendants.extend(self.get_user_descendants(child.id))

        return descendants


class ViewReferralStringAPIView(APIView):
    def get(self, request, user_id):
        user = CustomUser.objects.get(id=user_id)
        serializer = ReferalStringSerializer(user)
        return Response(serializer.data)

    def put(self, request, user_id):
        user = CustomUser.objects.get(id=user_id)
        serializer = ReferalStringSerializer(user, data=request.data)
        if serializer.is_valid():
            refer = "".join(random.choices(string.ascii_uppercase + string.digits, k=8))
            serializer.validated_data["referral_string"] = refer
            print(refer)
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# class EditReferralStringAPIView(APIView):


# def image_upload(request):
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_image(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    bearer_token = get_credential[0]["permanent_access_token"]

    bearer_token = urllib.parse.quote(bearer_token, safe="")
    url1 = (
        "https://graph.facebook.com/v18.0/329306953144911/uploads?access_token="
        + bearer_token
        + "&file_length=0&file_type=image/png"
    )

    response1 = requests.post(url1)

    if response1.status_code != 200:
        print(response1.text)  # Print the error response
        return JsonResponse(
            {"error": "Failed to make the first POST request"}, status=500
        )

    try:
        response1_data = response1.json()
        print(response1_data)
        upload_session_key = response1_data["id"]
    except ValueError:
        return JsonResponse({"error": "Invalid JSON in the first response"}, status=500)

    serializer = ImageUploadSerializer(data=request.data)
    if serializer.is_valid():
        image_file = serializer.validated_data["template_image"]
        template_name = serializer.validated_data["template_name"]
        content = image_file.read()

        file_name = os.path.basename(image_file.name)

        encoded_string = base64.b64encode(content).decode("utf-8")

        data_url = f"data:image;base64,{encoded_string}"

        url2 = "https://graph.facebook.com/v18.0/" + upload_session_key
        headers = {"Authorization": "OAuth " + bearer_token, "file_offset": "0"}

        # files = {"source": (file_name, open(file_name, "rb"))}
        files = {"source": data_url}
        response2 = requests.post(url2, headers=headers, files=files)
        serializer.save(user=request.user)

        if response2.status_code != 200:
            return JsonResponse({"error": response2.json()}, status=500)

        try:
            response2_data = response2.json()
            print(response2_data)
            return JsonResponse(response2_data)
        except ValueError:
            return JsonResponse(
                {"error": "Invalid JSON in the second response"}, status=500
            )

    return JsonResponse({"error": "Invalid data"}, status=400)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def delete_template(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]
    print(user_id)

    try:
        template_name = request.GET.get("template_name")
        cleaned_string = template_name.replace('"', "")

        url = (
            "https://graph.facebook.com/v18.0/"
            + business_id
            + "/message_templates?name="
            + cleaned_string
        )
        print(url)
        headers = {
            "Authorization": "Bearer " + bearer_token,
            "Content-Type": "application/json",
        }
        response = requests.request("DELETE", url, headers=headers)
        response_data = response.json()
        return Response(
            {"message": response_data},
            status=status.HTTP_201_CREATED,
        )

    except:
        return HttpResponse("ERROR")


def index(request):
    return HttpResponse("OK")


class PhoneNumberList(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        user_id = request.query_params.get("user_id")
        if user_id:
            phone_numbers = PhoneNumber.objects.filter(user__id=user_id)
            serializer = PhoneNumberSerializer(phone_numbers, many=True)
            numbers = [item["number"] for item in serializer.data]
        return Response(numbers, status=status.HTTP_200_OK)


# numbers upload from excel
class PhoneNumberUpload(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        serializer = ExcelSerializer(data=request.data)

        if serializer.is_valid():
            excel_file = serializer.validated_data["excel_file"]
            user_id = serializer.validated_data["user_id"]

            workbook = load_workbook(excel_file, read_only=True)
            worksheet = workbook.active

            for row in worksheet.iter_rows(values_only=True):
                raw_number = str(row[0])
                # print(raw_number)
                # replace 0 and add +91
                if raw_number.startswith("0"):
                    raw_number = "+91" + raw_number[1:]
                    # print(raw_number)
                # if no +91 add +91
                if not raw_number.startswith("+91"):
                    raw_number = "+91" + raw_number
                # model save
                PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

            return Response(
                {"message": "Phone numbers imported successfully"},
                status=status.HTTP_201_CREATED,
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def excel_sent_message(request):
    try:
        serializer = ExcelSerializer(data=request.data)

        if serializer.is_valid():
            excel_file = serializer.validated_data["excel_file"]
            template_name = serializer.validated_data["template_name"]
            user_id = serializer.validated_data["user_id"]
            get_credential = get_credentials(user_id)
            phone_number_id = get_credential[0]["phone_number_id"]
            # business_id = get_credential[0]["whatsapp_business_id"]
            bearer_token = get_credential[0]["permanent_access_token"]

            workbook = load_workbook(excel_file, read_only=True)
            worksheet = workbook.active
            results = []

            for row in worksheet.iter_rows(values_only=True):
                raw_number = str(row[0])
                print(raw_number)
                # replace 0 and add +91
                if raw_number.startswith("0"):
                    raw_number = "+91" + raw_number[1:]
                # if no +91 add +91
                if not raw_number.startswith("+91"):
                    raw_number = "+91" + raw_number

                PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

                data = {
                    "messaging_product": "whatsapp",
                    "recipient_type": "individual",
                    "to": raw_number,
                    "type": "template",
                    "template": {"name": template_name, "language": {"code": "en"}},
                }
                # print(data)

                url = (
                    "https://graph.facebook.com/v18.0/" + phone_number_id + "/messages"
                )
                headers = {
                    "Authorization": "Bearer " + bearer_token,
                    "Content-Type": "application/json",
                }

                try:
                    response = requests.post(url, headers=headers, json=data)
                    response_data = response.json()
                    results.append(response_data)

                except json.JSONDecodeError:
                    return JsonResponse({"error": "Invalid JSON data"}, status=400)
            return Response(
                {"message": "Phone numbers imported successfully"},
                status=status.HTTP_201_CREATED,
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def excel_personalised_sent_message(request):
    try:
        serializer = ExcelSerializer(data=request.data)

        if serializer.is_valid():
            excel_file = serializer.validated_data["excel_file"]
            template_name = serializer.validated_data["template_name"]
            user_id = serializer.validated_data["user_id"]
            get_credential = get_credentials(user_id)
            phone_number_id = get_credential[0]["phone_number_id"]
            # business_id = get_credential[0]["whatsapp_business_id"]
            bearer_token = get_credential[0]["permanent_access_token"]

            workbook = load_workbook(excel_file, read_only=True)
            worksheet = workbook.active
            results = []

            for row in worksheet.iter_rows(values_only=True):
                raw_number = str(row[1])
                name = str(row[0])
                print(name, raw_number)
                # replace 0 and add +91
                if raw_number.startswith("0"):
                    raw_number = "+91" + raw_number[1:]
                # if no +91 add +91
                if not raw_number.startswith("+91"):
                    raw_number = "+91" + raw_number

                print(raw_number)
                # model save
                # PhoneNumber.objects.get_or_create(number=raw_number)
                PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

                data = {
                    "messaging_product": "whatsapp",
                    "recipient_type": "individual",
                    "to": raw_number,
                    "type": "template",
                    "template": {
                        "name": template_name,
                        "components": [
                            {
                                "type": "HEADER",
                                # "format": "TEXT",
                                "parameters": [{"type": "text", "text": name}]
                                # "example": {"header_text": ["Jeevan"]},
                            }
                        ],
                        "language": {"code": "en"},
                    },
                }

                url = (
                    "https://graph.facebook.com/v18.0/" + phone_number_id + "/messages"
                )
                headers = {
                    "Authorization": "Bearer " + bearer_token,
                    "Content-Type": "application/json",
                }

                try:
                    response = requests.post(url, headers=headers, json=data)
                    response_data = response.json()
                    results.append(response_data)

                except json.JSONDecodeError:
                    return JsonResponse({"error": "Invalid JSON data"}, status=400)
            return Response(
                {"message": response_data},
                status=status.HTTP_201_CREATED,
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def excel_sent_message_images(request):
    try:
        serializer = ExcelImageSerializer(data=request.data)

        if serializer.is_valid():
            excel_file = serializer.validated_data["excel_file"]
            template_name = serializer.validated_data["template_name"]
            user_id = serializer.validated_data["user_id"]
            image_link = serializer.validated_data["image_link"]
            get_credential = get_credentials(user_id)
            phone_number_id = get_credential[0]["phone_number_id"]
            # business_id = get_credential[0]["whatsapp_business_id"]
            bearer_token = get_credential[0]["permanent_access_token"]

            workbook = load_workbook(excel_file, read_only=True)
            worksheet = workbook.active
            results = []

            for row in worksheet.iter_rows(values_only=True):
                raw_number = str(row[0])
                print(raw_number)

                # replace 0 and add +91
                if raw_number.startswith("0"):
                    raw_number = "+91" + raw_number[1:]
                # if no +91 add +91
                if not raw_number.startswith("+91"):
                    raw_number = "+91" + raw_number
                # model save
                # PhoneNumber.objects.get_or_create(number=raw_number)
                PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

                data = {
                    "messaging_product": "whatsapp",
                    "recipient_type": "individual",
                    "to": raw_number,
                    "type": "template",
                    "template": {
                        "name": template_name,
                        "components": [
                            {
                                "type": "HEADER",
                                # "format": "IMAGE",
                                "parameters": [
                                    {
                                        "type": "image",
                                        "image": {
                                            "link": "https://i.ibb.co/23hxBhg/image.png"
                                        },
                                    }
                                ],
                            }
                        ],
                        "language": {"code": "en"},
                    },
                }
                # print(data)

                url = (
                    "https://graph.facebook.com/v18.0/" + phone_number_id + "/messages"
                )
                headers = {
                    "Authorization": "Bearer " + bearer_token,
                    "Content-Type": "application/json",
                }

                try:
                    response = requests.post(url, headers=headers, json=data)
                    response_data = response.json()
                    results.append(response_data)

                except json.JSONDecodeError:
                    return JsonResponse({"error": "Invalid JSON data"}, status=400)
            return Response(
                {"message": results},
                status=status.HTTP_201_CREATED,
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def excel_sent_message_images_personalised(request):
    try:
        serializer = ExcelImageSerializer(data=request.data)

        if serializer.is_valid():
            excel_file = serializer.validated_data["excel_file"]
            template_name = serializer.validated_data["template_name"]
            user_id = serializer.validated_data["user_id"]
            image_link = serializer.validated_data["image_link"]
            get_credential = get_credentials(user_id)
            phone_number_id = get_credential[0]["phone_number_id"]
            # business_id = get_credential[0]["whatsapp_business_id"]
            bearer_token = get_credential[0]["permanent_access_token"]

            workbook = load_workbook(excel_file, read_only=True)
            worksheet = workbook.active
            results = []

            for row in worksheet.iter_rows(values_only=True):
                raw_number = str(row[1])
                name = str(row[0])
                print(name)
                # replace 0 and add +91
                if raw_number.startswith("0"):
                    raw_number = "+91" + raw_number[1:]
                # if no +91 add +91
                if not raw_number.startswith("+91"):
                    raw_number = "+91" + raw_number

                print(raw_number)
                # model save
                # PhoneNumber.objects.get_or_create(number=raw_number)
                # PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

                # data = {
                #     "messaging_product": "whatsapp",
                #     "recipient_type": "individual",
                #     "to": raw_number,
                #     "type": "template",
                #     "template": {
                #         "name": template_name,
                #         "components": [
                #             {
                #                 "type": "HEADER",
                #                 # "format": "IMAGE",
                #                 "parameters": [
                #                     {
                #                         "type": "image",
                #                         "image": {
                #                             # "link": image_link
                #                             "link": "https://i.ibb.co/GTmcbDg/386384270-632166208949549-651046045893776904-n.png"
                #                         },
                #                     }
                #                 ],
                #             },
                #             {
                #                 "type": "body",
                #                 "parameters": [{"type": "text", "text": name}],
                #             },
                #         ],
                #         "language": {"code": "en"},
                #     },
                # }
                # # print(data)

                # url = (
                #     "https://graph.facebook.com/v18.0/" + phone_number_id + "/messages"
                # )
                # headers = {
                #     "Authorization": "Bearer " + bearer_token,
                #     "Content-Type": "application/json",
                # }

                # try:
                #     response = requests.post(url, headers=headers, json=data)
                #     response_data = response.json()
                #     results.append(response_data)

                # except json.JSONDecodeError:
                #     return JsonResponse({"error": "Invalid JSON data"}, status=400)
            return Response(
                {"message": results},
                status=status.HTTP_201_CREATED,
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def excel_upload_message(request):
    try:
        serializer = ExcelSerializer(data=request.data)

        if serializer.is_valid():
            print("raw_number")

            excel_file = serializer.validated_data["excel_file"]
            user_id = serializer.validated_data["user_id"]

            workbook = load_workbook(excel_file, read_only=True)
            worksheet = workbook.active
            results = []

            for row in worksheet.iter_rows(values_only=True):
                raw_number = str(row[0])
                print(raw_number)
                print("raw_number")
                # replace 0 and add +91
                if raw_number.startswith("0"):
                    raw_number = "+91" + raw_number[1:]
                # if no +91 add +91
                if not raw_number.startswith("+91"):
                    raw_number = "+91" + raw_number
                # model save
                # PhoneNumber.objects.get_or_create(number=raw_number)
                # print(raw_number)
                PhoneNumber.objects.get_or_create(number=raw_number, user_id=user_id)

            return Response(
                {"message": "Phone numbers imported successfully"},
                status=status.HTTP_201_CREATED,
            )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def send_whatsapp_bulk_messages(request):
    try:
        serializer = WhatsAppBulkMessageSerializer(data=request.data)

        if serializer.is_valid():
            template_name = serializer.validated_data.get("template_name")
            numbers = serializer.validated_data.get("numbers")
            user_id = serializer.validated_data.get("user_id")
            get_credential = get_credentials(user_id)
            print(get_credential)
            phone_number_id = get_credential[0]["phone_number_id"]
            # business_id = get_credential[0]["whatsapp_business_id"]
            bearer_token = get_credential[0]["permanent_access_token"]
            results = []

        for number in numbers:
            if number.startswith("0"):
                number = "+91" + number[1:]

            if not number.startswith("+91"):
                number = "+91" + number

            if not number:
                results.append({"error": "Missing 'number' parameter"})
                continue
            PhoneNumber.objects.get_or_create(number=number, user_id=user_id)

            data = {
                "messaging_product": "whatsapp",
                "recipient_type": "individual",
                "to": number,
                "type": "template",
                "template": {"name": template_name, "language": {"code": "en"}},
            }
            url = "https://graph.facebook.com/v18.0/" + phone_number_id + "/messages"
            headers = {
                "Authorization": "Bearer " + bearer_token,
                "Content-Type": "application/json",
            }

            try:
                response = requests.post(url, headers=headers, json=data)
                response_data = response.json()
                results.append(response_data)
            except requests.exceptions.RequestException as e:
                results.append({"error": str(e)})
            except json.JSONDecodeError:
                results.append({"error": "Invalid JSON data"})

        for i, result in enumerate(results):
            if not isinstance(result, dict):
                results[i] = {"error": str(result)}

        return JsonResponse(results, status=200, safe=False)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON data"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# image
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def send_whatsapp_bulk_messages_images(request):
    try:
        serializer = WhatsAppBulkMessageImageSerializer(data=request.data)

        if serializer.is_valid():
            template_name = serializer.validated_data.get("template_name")
            numbers = serializer.validated_data.get("numbers")
            image_link = serializer.validated_data.get("image_link")
            user_id = serializer.validated_data.get("user_id")
            get_credential = get_credentials(user_id)
            phone_number_id = get_credential[0]["phone_number_id"]
            # business_id = get_credential[0]["whatsapp_business_id"]
            bearer_token = get_credential[0]["permanent_access_token"]

            # print(template_name)
            results = []

        for number in numbers:
            if number.startswith("0"):
                number = "+91" + number[1:]

            if not number.startswith("+91"):
                number = "+91" + number

            if not number:
                results.append({"error": "Missing 'number' parameter"})
                continue
            PhoneNumber.objects.get_or_create(number=number, user_id=user_id)

            data = {
                "messaging_product": "whatsapp",
                "recipient_type": "individual",
                "to": number,
                "type": "template",
                "template": {
                    "name": template_name,
                    "components": [
                        # {
                        #     "type": "HEADER",
                        #     # "format": "TEXT",
                        #     "parameters": [{"type": "text", "text": "MANU"}]
                        #     # "example": {"header_text": ["Jeevan"]},
                        # }
                        {
                            "type": "header",
                            # "format": "IMAGE",
                            # "example": {"header_handle": image_link},
                            "parameters": [
                                {
                                    "type": "image",
                                    "image": {
                                        # "link": image_link
                                        # "link": "https://i.ibb.co/QcpC8WQ/bgnotfound.png"
                                        "link": "https://i.ibb.co/23hxBhg/image.png"
                                    },
                                }
                            ],
                        }
                    ],
                    "language": {"code": "en"},
                },
            }
            url = "https://graph.facebook.com/v18.0/" + phone_number_id + "/messages"
            headers = {
                "Authorization": "Bearer " + bearer_token,
                "Content-Type": "application/json",
            }

            try:
                response = requests.post(url, headers=headers, json=data)
                response_data = response.json()
                results.append(response_data)
            except requests.exceptions.RequestException as e:
                results.append({"error": str(e)})
            except json.JSONDecodeError:
                results.append({"error": "Invalid JSON data"})

        for i, result in enumerate(results):
            if not isinstance(result, dict):
                results[i] = {"error": str(result)}

        return JsonResponse(results, status=200, safe=False)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON data"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# send bulk messages from addded names from database
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def send_whatsapp_model_bulk_messages(request):
    try:
        template_name = request.GET.get("template_name")
        user_id = request.GET.get("user_id")
        get_credential = get_credentials(user_id)
        phone_number_id = get_credential[0]["phone_number_id"]
        # business_id = get_credential[0]["whatsapp_business_id"]
        bearer_token = get_credential[0]["permanent_access_token"]
        user_instance = CustomUser.objects.get(pk=user_id)
        get_numbers = PhoneNumber.objects.filter(user_id=user_id).values_list(
            "number", flat=True
        )
        numbers = list(get_numbers)
        print(get_numbers)

        if not numbers or not isinstance(numbers, list):
            return JsonResponse(
                {"error": "Missing or invalid 'numbers' parameter"}, status=400
            )
        results = []

        for number in numbers:
            if not number:
                results.append({"error": "Missing 'number' parameter"})
                continue

            data = {
                "messaging_product": "whatsapp",
                "recipient_type": "individual",
                "to": number,
                "type": "template",
                "template": {"name": template_name, "language": {"code": "en"}},
            }

            url = "https://graph.facebook.com/v18.0/" + phone_number_id + "/messages"
            headers = {
                "Authorization": "Bearer " + bearer_token,
                "Content-Type": "application/json",
            }

            try:
                response = requests.post(url, headers=headers, json=data)
                response_data = response.json()
                results.append(response_data)
            except requests.exceptions.RequestException as e:
                results.append({"error": str(e)})
            except json.JSONDecodeError:
                results.append({"error": "Invalid JSON data"})

        for i, result in enumerate(results):
            if not isinstance(result, dict):
                results[i] = {"error": str(result)}

        return JsonResponse(results, status=200, safe=False)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON data"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# send bulk messages from addded names from database
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def send_whatsapp_model_bulk_messages_images(request):
    try:
        template_name = request.GET.get("template_name")
        user_id = request.GET.get("user_id")
        get_credential = get_credentials(user_id)
        phone_number_id = get_credential[0]["phone_number_id"]
        # business_id = get_credential[0]["whatsapp_business_id"]
        bearer_token = get_credential[0]["permanent_access_token"]
        image_url = request.GET.get("image_url")
        get_numbers = PhoneNumber.objects.filter(user_id=user_id).values_list(
            "number", flat=True
        )
        numbers = list(get_numbers)
        template_instance = get_object_or_404(
            Template, user__id=user_id, template_name=template_name
        )

        # print(f"{domain_url}{template_instance.template_image}")

        if not numbers or not isinstance(numbers, list):
            return JsonResponse(
                {"error": "Missing or invalid 'numbers' parameter"}, status=400
            )
        results = []

        for number in numbers:
            if not number:
                results.append({"error": "Missing 'number' parameter"})
                continue

            data = {
                "messaging_product": "whatsapp",
                "recipient_type": "individual",
                "to": number,
                "type": "template",
                "template": {
                    "name": template_name,
                    "components": [
                        {
                            "type": "HEADER",
                            # "format": "IMAGE",
                            "parameters": [
                                {
                                    "type": "image",
                                    "image": {
                                        "link": "https://i.ibb.co/23hxBhg/image.png"
                                    },
                                }
                            ],
                        }
                    ],
                    "language": {"code": "en"},
                },
            }
            # print(data)

            url = "https://graph.facebook.com/v18.0/" + phone_number_id + "/messages"
            headers = {
                "Authorization": "Bearer " + bearer_token,
                "Content-Type": "application/json",
            }

            try:
                response = requests.post(url, headers=headers, json=data)
                response_data = response.json()
                results.append(response_data)
            except requests.exceptions.RequestException as e:
                results.append({"error": str(e)})
            except json.JSONDecodeError:
                results.append({"error": "Invalid JSON data"})

        for i, result in enumerate(results):
            if not isinstance(result, dict):
                results[i] = {"error": str(result)}

        return JsonResponse(results, status=200, safe=False)

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON data"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# get templates
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_templates_message(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]
    template_instances = Template.objects.filter(user__id=user_id)

    template_image_array = []

    for template_instance in template_instances:
        template_dict = {
            template_instance.template_name: template_instance.template_image.url
        }
        template_image_array.append(template_dict)

    print(template_image_array)

    url = "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    headers = {
        "Authorization": "Bearer " + bearer_token,
    }
    print(user_id)

    try:
        response = requests.get(url, headers=headers)
        data = response.json()
        templates = data.get("data", [])

        names = [template.get("name", "") for template in templates]
        components = [template.get("components", []) for template in templates]

        name_response = {
            "names": names,
            "components": components,
            "images": template_image_array,
        }

        # print(data)
        return JsonResponse({"data": name_response})
        # return JsonResponse({"data": templates})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_templates_list(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]

    url = "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    headers = {
        "Authorization": "Bearer " + bearer_token,
    }

    try:
        response = requests.get(url, headers=headers)
        data = response.json()
        templates = data.get("data", [])

        names = [template.get("name", "") for template in templates]
        components = [template.get("components", []) for template in templates]

        name_response = {"names": names, "components": components}

        # print(data)
        return JsonResponse({"data": templates})
        # return JsonResponse({"data": name_response})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_text_template(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]
    facebook_api_url = (
        "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    )
    serializer = MessageTextTemplateSerializer(data=request.data)

    if serializer.is_valid():
        data = serializer.validated_data
        template_name = data.get("template_name")
        header_text = data.get("header_text")
        body_text = data.get("body_text")
        footer_text = data.get("footer_text")

        post_data = json.dumps(
            {
                "name": template_name,
                "category": "MARKETING",
                "language": "en",
                "components": [
                    # {
                    #     "type": "HEADER",
                    #     "format": "TEXT",
                    #     "text": "Our {{1}} is on!",
                    #     "example": {"header_text": ["Sample Text"]},
                    # },
                    {
                        "type": "HEADER",
                        "format": "TEXT",
                        "text": header_text,
                    },
                    {"type": "BODY", "text": body_text},
                    {"type": "FOOTER", "text": footer_text},
                ],
            }
        )

        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + bearer_token,
        }

        response = requests.post(facebook_api_url, data=post_data, headers=headers)
        response_data = response.json()

        if response.status_code == status.HTTP_200_OK:
            return Response(
                {"message": response_data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"message": response_data},
                status=response.status_code,
            )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_image_template(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]

    facebook_api_url = (
        "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    )
    serializer = MessageTemplateSerializer(data=request.data)

    if serializer.is_valid():
        data = serializer.validated_data
        template_name = data.get("template_name")
        header_text = data.get("header_text")
        body_text = data.get("body_text")
        footer_text = data.get("footer_text")

        print(header_text)

        post_data = json.dumps(
            {
                "name": template_name,
                "category": "MARKETING",
                "language": "en",
                "components": [
                    {
                        "type": "HEADER",
                        "format": "IMAGE",
                        "example": {"header_handle": header_text},
                    },
                    {"type": "BODY", "text": body_text},
                    {"type": "FOOTER", "text": footer_text},
                ],
            }
        )

        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + bearer_token,
        }

        response = requests.post(facebook_api_url, data=post_data, headers=headers)
        response_data = response.json()

        if response.status_code == status.HTTP_200_OK:
            return Response(
                {"message": response_data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"message": response_data},
                status=response.status_code,
            )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_image_template_personalised(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]

    facebook_api_url = (
        "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    )
    serializer = MessageTemplateSerializer(data=request.data)

    if serializer.is_valid():
        data = serializer.validated_data
        template_name = data.get("template_name")
        header_text = data.get("header_text")
        body_text = data.get("body_text")
        footer_text = data.get("footer_text")
        print(body_text)
        print(header_text)

        post_data = json.dumps(
            {
                "name": template_name,
                "category": "MARKETING",
                "language": "en",
                "components": [
                    {
                        "type": "HEADER",
                        "format": "IMAGE",
                        "example": {"header_handle": header_text},
                    },
                    {
                        "type": "BODY",
                        # "text": "Thank you for your order, {{1}}! Your confirmation number is {{2}}. If you have any questions, please use the buttons below to contact support. Thank you for being a customer!",
                        # "example": {"body_text": [["Pablo", "860198-230332"]]}
                        "text": f"{body_text}",
                        "example": {"body_text": [["Sample"]]},
                    },
                    {"type": "FOOTER", "text": footer_text},
                ],
            }
        )

        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + bearer_token,
        }

        response = requests.post(facebook_api_url, data=post_data, headers=headers)
        response_data = response.json()

        if response.status_code == status.HTTP_200_OK:
            return Response(
                {"message": response_data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"message": response_data},
                status=response.status_code,
            )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_text_template_button_site(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]

    facebook_api_url = (
        "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    )
    serializer = MessageTemplateSerializer(data=request.data)

    if serializer.is_valid():
        data = serializer.validated_data
        template_name = data.get("template_name")
        header_text = data.get("header_text")
        body_text = data.get("body_text")
        footer_text = data.get("footer_text")
        button_type = data.get("button_type")
        button_text = data.get("button_text")
        button_url = data.get("button_url")

        post_data = json.dumps(
            {
                "name": template_name,
                "category": "MARKETING",
                "components": [
                    {
                        "type": "HEADER",
                        "format": "TEXT",
                        "text": header_text,
                    },
                    {
                        "type": "BODY",
                        "text": body_text,
                    },
                    {"type": "FOOTER", "text": footer_text},
                    {
                        "type": "BUTTONS",
                        "buttons": [
                            {
                                "type": "URL",
                                "text": button_text,
                                "url": button_url,
                            }
                        ],
                    },
                ],
                "language": "en",
            }
        )

        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + bearer_token,
        }

        response = requests.post(facebook_api_url, data=post_data, headers=headers)
        response_data = response.json()

        if response.status_code == status.HTTP_200_OK:
            return Response(
                {"message": response_data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"message": response_data},
                status=response.status_code,
            )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_text_template_button_call(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]

    facebook_api_url = (
        "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    )
    serializer = MessageTemplateSerializer(data=request.data)

    if serializer.is_valid():
        data = serializer.validated_data
        template_name = data.get("template_name")
        header_text = data.get("header_text")
        body_text = data.get("body_text")
        footer_text = data.get("footer_text")
        button_type = data.get("button_type")
        button_text = data.get("button_text")
        button_url = data.get("button_url")

        post_data = json.dumps(
            {
                "name": template_name,
                "category": "MARKETING",
                "components": [
                    {"type": "HEADER", "format": "TEXT", "text": header_text},
                    {
                        "type": "BODY",
                        "text": body_text,
                    },
                    {"type": "FOOTER", "text": footer_text},
                    {
                        "type": "BUTTONS",
                        "buttons": [
                            {
                                "type": "PHONE_NUMBER",
                                "text": button_text,
                                "phone_number": button_url,
                            }
                        ],
                    },
                ],
                "language": "en",
            }
        )

        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + bearer_token,
        }

        response = requests.post(facebook_api_url, data=post_data, headers=headers)
        response_data = response.json()

        if response.status_code == status.HTTP_200_OK:
            return Response(
                {"message": response_data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"message": response_data},
                status=response.status_code,
            )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_text_template_personalised(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]
    facebook_api_url = (
        "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    )
    serializer = MessageTextTemplateSerializer(data=request.data)

    if serializer.is_valid():
        data = serializer.validated_data
        template_name = data.get("template_name")
        header_text = data.get("header_text")
        body_text = data.get("body_text")
        footer_text = data.get("footer_text")

        post_data = json.dumps(
            {
                "name": template_name,
                "category": "MARKETING",
                "language": "en",
                "components": [
                    {
                        "type": "HEADER",
                        "format": "TEXT",
                        "text": header_text,
                        "example": {"header_text": ["Sample Text"]},
                    },
                    # {"type": "HEADER", "format": "TEXT", "text": header_text},
                    {"type": "BODY", "text": body_text},
                    {"type": "FOOTER", "text": footer_text},
                ],
            }
        )

        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + bearer_token,
        }

        response = requests.post(facebook_api_url, data=post_data, headers=headers)
        response_data = response.json()

        if response.status_code == status.HTTP_200_OK:
            return Response(
                {"message": response_data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"message": response_data},
                status=response.status_code,
            )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_text_template_button_site_personalised(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]

    facebook_api_url = (
        "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    )
    serializer = MessageTemplateSerializer(data=request.data)

    if serializer.is_valid():
        data = serializer.validated_data
        template_name = data.get("template_name")
        header_text = data.get("header_text")
        body_text = data.get("body_text")
        footer_text = data.get("footer_text")
        button_type = data.get("button_type")
        button_text = data.get("button_text")
        button_url = data.get("button_url")

        post_data = json.dumps(
            {
                "name": template_name,
                "category": "MARKETING",
                "components": [
                    {
                        "type": "HEADER",
                        "format": "TEXT",
                        "text": header_text,
                        "example": {"header_text": ["Sample Text"]},
                    },
                    {
                        "type": "BODY",
                        "text": body_text,
                    },
                    {"type": "FOOTER", "text": footer_text},
                    {
                        "type": "BUTTONS",
                        "buttons": [
                            {
                                "type": "URL",
                                "text": button_text,
                                "url": button_url,
                            }
                        ],
                    },
                ],
                "language": "en",
            }
        )

        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + bearer_token,
        }

        response = requests.post(facebook_api_url, data=post_data, headers=headers)
        response_data = response.json()

        if response.status_code == status.HTTP_200_OK:
            return Response(
                {"message": response_data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"message": response_data},
                status=response.status_code,
            )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_text_template_button_call_personalised(request):
    user_id = request.GET.get("user_id")
    get_credential = get_credentials(user_id)
    # phone_number_id = get_credential[0]["phone_number_id"]
    business_id = get_credential[0]["whatsapp_business_id"]
    bearer_token = get_credential[0]["permanent_access_token"]

    facebook_api_url = (
        "https://graph.facebook.com/v18.0/" + business_id + "/message_templates"
    )
    serializer = MessageTemplateSerializer(data=request.data)

    if serializer.is_valid():
        data = serializer.validated_data
        template_name = data.get("template_name")
        header_text = data.get("header_text")
        body_text = data.get("body_text")
        footer_text = data.get("footer_text")
        button_type = data.get("button_type")
        button_text = data.get("button_text")
        button_url = data.get("button_url")

        post_data = json.dumps(
            {
                "name": template_name,
                "category": "MARKETING",
                "components": [
                    {
                        "type": "HEADER",
                        "format": "TEXT",
                        "text": header_text,
                        "example": {"header_text": ["Sample Text"]},
                    },
                    {
                        "type": "BODY",
                        "text": body_text,
                    },
                    {"type": "FOOTER", "text": footer_text},
                    {
                        "type": "BUTTONS",
                        "buttons": [
                            {
                                "type": "PHONE_NUMBER",
                                "text": button_text,
                                "phone_number": button_url,
                            }
                        ],
                    },
                ],
                "language": "en",
            }
        )

        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + bearer_token,
        }

        response = requests.post(facebook_api_url, data=post_data, headers=headers)
        response_data = response.json()

        if response.status_code == status.HTTP_200_OK:
            return Response(
                {"message": response_data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"message": response_data},
                status=response.status_code,
            )
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# privacy policy
def privacy(request):
    return JsonResponse(
        {
            "privacy & Policy": "Message FlowsWhen a user sends a message to one of these businesses, the message travels end-to-end encrypted between the user and the Cloud API. As per the Signal protocol, the user and the Cloud API, on behalf of the business, negotiate encryption keys and establish a secure communication channel. WhatsApp cannot access any message content exchanged between users and businesses.Once a message is received by the Cloud API, it gets decrypted and forwarded to the Business. Messages are only temporarily stored by the Cloud API as required to provide the base API functionality.Messages from a business to a user flow on the reverse path. Businesses send messages to Cloud API. The Cloud API service stores the messages temporarily and takes on the task to send the message to the WhatsApp platform. Messages are stored for any necessary retransmissions.All messages are encrypted by the Cloud API before being sent to WhatsApp using the Signal protocol with keys negotiated with the user (recipient).WhatsApp acts as the transport service. It provides the message forwarding software; both client and server. It has no visibility into the messages being sent. It protects the users by detecting unusual messaging patterns (like a business trying to message all users) or collecting spam reports from users.Cloud API, operated by Meta, acts as the intermediary between WhatsApp and the Cloud API businesses. In other words, those businesses have given Cloud API the power to operate on their behalf. Because of this, WhatsApp forwards all message traffic destined for those businesses to Cloud API. WhatsApp also expects to receive from Cloud API all message traffic from those businesses. This is the same client behavior that the On-Premise client has.WhatsApp gives Cloud API metering and billing information for the Cloud API businesses. It does not share any other messaging information.Meta, in providing the WhatsApp Cloud API service, acts as a Data Processor on behalf of the business. In other words, the businesses have requested Meta to provide programmatic access to the WhatsApp platform.Cloud API receives from WhatsApp the messages destined for the businesses that use Cloud API. Cloud API also sends to WhatsApp the messages sent by those businesses. Other parts of Meta (other than Cloud API) do not have access to the Cloud API business communications, including message content and metadata. Meta does not use any Cloud API data for advertising.Stored and Collected DataAll data collected, stored and accessed by Cloud API is controlled and monitored to ensure proper usage and maintain the high level of privacy expected from a WhatsApp client.Information about the businesses, including their phone numbers, business address, contacts, type, etc. is maintained by Meta and the Business Manager product and is subject to the terms of service set by Meta. Cloud API relies on Business Manager and other Meta systems to identify any access to Cloud API on behalf of the business.Messages sent or received via Cloud API are only accessed by Cloud API, no other part of Meta can use this information. Messages have a maximum retention period of 30 days in order to provide the base features and functionality of the Cloud API service; for example, retransmissions. After 30 days, these features and functionality are no longer available.Cloud API does not rely on any information about the user (customer/consumer) the business is communicating with other than the phone number used to identify the account. This information is used to deliver the messages via the WhatsApp client code. User phone numbers are used as sources or destinations of individual messages; as such they are deleted when messages are deleted. No other part of Meta has access to this information. Like the On-Premise client, the WhatsApp client code used by Cloud API collects messaging information about the business as required by WhatsApp. This is information used by WhatsApp to detect malicious activity. No message content is shared or sent to WhatsApp at any time and no WhatsApp employee has access to any message content."
        }
    )


# webhook handle
@csrf_exempt
def whatsapp_webhook(request):
    if request.method == "GET":
        mode = request.GET.get("hub.mode")
        challenge = request.GET.get("hub.challenge")
        token = request.GET.get("hub.verify_token")

        if mode and token:
            if mode == "subscribe" and token == my_token:
                return HttpResponse(challenge, content_type="text/plain")
            else:
                return HttpResponse(status=403)

    return HttpResponse("OK", content_type="text/plain")


from datetime import datetime, timedelta, timezone

# from .tasks import make_api_call

import pytz


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from datetime import datetime
from .tasks import schedule_hello


class ScheduleHelloView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            # Get the datetime parameter from the request data
            datetime_param = request.data.get("datetime_param")
            print(timezone.now())

            # Parse the datetime string into a datetime object
            target_datetime = timezone.datetime.strptime(
                datetime_param, "%Y-%m-%dT%H:%M:%S.%fZ"
            )

            # Schedule the task to print "hello" at the specified datetime
            schedule_hello.apply_async((target_datetime,), eta=target_datetime)

            return Response(
                {"message": "Task scheduled successfully"}, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


@csrf_exempt
@api_view(["POST"])
def schedule_api_call(request):
    print("Request Data:", request.data)

    scheduled_time_str = request.data.get("scheduled_time")
    api_data = request.data.get("api_data")

    if scheduled_time_str is not None:
        try:
            time = datetime.now(timezone.utc)
            current_time_utc = datetime.utcnow().replace(tzinfo=pytz.UTC)

            # Convert UTC time to IST
            ist = pytz.timezone("Asia/Kolkata")
            print(time)
            current_time_ist = current_time_utc.astimezone(ist)
            current_time_ist_plus_one_minute = current_time_ist + timedelta(minutes=1)
            print(current_time_ist)
            print(current_time_ist_plus_one_minute)

            task = ScheduledAPICall.objects.create(
                api_data=api_data, scheduled_time=time
            )

            make_api_call.apply_async(
                args=[task.id], eta=current_time_ist_plus_one_minute
            )

            return HttpResponse("API call scheduled successfully!")
        except ValueError as e:
            return HttpResponse(f"Error: {e}", status=400)
    else:
        return HttpResponse("Error: Scheduled time is required", status=400)


# @csrf_exempt
# def whatsapp_message(request):
#     if request.method == "POST":
#         print("sdhjbcjhdc", request.body.decode("utf-8"))
#         print(json.loads(request.body.decode("utf-8")))
#         try:
#             body_param = json.loads(request.body.decode("utf-8"))

#             if "object" in body_param:
#                 entry = body_param.get("entry")
#                 if (
#                     entry
#                     and entry[0].get("changes")
#                     and entry[0]["changes"][0].get("value")
#                     and entry[0]["changes"][0]["value"].get("messages")
#                 ):
#                     phone_number_id = entry[0]["changes"][0]["value"]["metadata"][
#                         "phone_number_id"
#                     ]
#                     from_user = entry[0]["changes"][0]["value"]["messages"][0]["from"]
#                     msg_body = entry[0]["changes"][0]["value"]["messages"][0]["text"][
#                         "body"
#                     ]

#                     print("phone number", phone_number_id)
#                     print("from", from_user)
#                     print("message body", msg_body)

#                     # Define your Facebook Graph API endpoint
#                     facebook_api_url = (
#                         "https://graph.facebook.com/v13.0/{phone_number_id}/messages"
#                     )

#                     # Create the message payload
#                     message_payload = {
#                         "messaging_product": "whatsapp",
#                         "to": from_user,
#                         "text": {
#                             "body": "Hi.. I'm Prasath, your message is {msg_body}"
#                         },
#                     }

#                     # Send the POST request to the Facebook Graph API
#                     headers = {"Content-Type": "application/json"}
#                     params = {"access_token": token}
#                     response = requests.post(
#                         facebook_api_url,
#                         json=message_payload,
#                         headers=headers,
#                         params=params,
#                     )

#                     if response.status_code == 200:
#                         return HttpResponse(status=200)
#                     else:
#                         return HttpResponse(status=500)
#         except json.JSONDecodeError as e:
#             return HttpResponse(
#                 "Invalid JSON data", content_type="text/plain", status=400
#             )


# sent individual message
# @csrf_exempt
# @require_POST
# def send_whatsapp_message(request):
#     try:
#         post_data = json.loads(request.body.decode("utf-8"))
#         to = post_data.get("name")
#         # print("Received 'to' parameter: ", to)

#         PhoneNumber.objects.get_or_create(number=to)

#         if not to:
#             return JsonResponse({"error": "Missing 'to' parameter"}, status=400)

#         data = {
#             "messaging_product": "whatsapp",
#             "recipient_type": "individual",
#             "to": to,
#             "type": "template",
#             "template": {"name": "developer_test", "language": {"code": "en"}},
#         }

#         url = "https://graph.facebook.com/v18.0/123004784227116/messages"
#         headers = {
#             "Authorization": "Bearer "
#             + bearer_token,  # Include the Bearer token in the headers
#             "Content-Type": "application/json",
#         }
#         # print(headers)

#         response = requests.post(url, headers=headers, json=data)
#         response_data = response.json()
#         return JsonResponse(response_data, status=response.status_code)

#     except json.JSONDecodeError:
#         return JsonResponse({"error": "Invalid JSON data"}, status=400)
#     except requests.exceptions.RequestException as e:
#         return JsonResponse({"error": str(e)}, status=500)

# from django.core.mail import send_mail, EmailMultiAlternatives
# from django.template.loader import render_to_string
# from django.utils.html import strip_tags
# from django.conf import settings
# from io import BytesIO
# from xhtml2pdf import pisa


# def pdf(request):
#     if request.method == "POST":
#         my_subject = "Whatsapp Module Generated Password"


#         recipient_list = ["jeevanjose2016@gmail.com"]#add ur email
#         html_message = render_to_string("password.html")#add ur html
#         plain_message = strip_tags(html_message)
#         pdf_content = BytesIO()
#         pisa_document = pisa.CreatePDF(
#             BytesIO(html_message.encode("UTF-8")), pdf_content
#         )
#         pdf_content.seek(0)
#         message = EmailMultiAlternatives(
#             subject=my_subject,
#             body="plain_message",
#             from_email=settings.DEFAULT_FROM_EMAIL,
#             to=recipient_list,
#         )
#         message.attach("filename", pdf_content.read(), "application/pdf")
#         message.send()

#         return HttpResponse("SUccess")
#     else:
#         return HttpResponse("ERROR")
